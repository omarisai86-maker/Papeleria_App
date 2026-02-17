from fastapi import FastAPI
from fastapi.middleware.cors import CORSMiddleware
from fastapi.staticfiles import StaticFiles
from fastapi.responses import FileResponse
import sqlite3
from openpyxl import load_workbook
import os

app = FastAPI()

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# =========================
# BASE DE DATOS
# =========================

def init_db():
    conn = sqlite3.connect("database.db")
    cursor = conn.cursor()

    cursor.execute("""
    CREATE TABLE IF NOT EXISTS productos (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        codigo TEXT,
        nombre TEXT,
        compra REAL,
        porcentaje REAL,
        venta REAL,
        cantidad_existente INTEGER,
        cantidad_comprar INTEGER
    )
    """)

    cursor.execute("""
    CREATE TABLE IF NOT EXISTS historial (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        codigo TEXT,
        nombre TEXT,
        compra REAL,
        porcentaje REAL,
        venta REAL,
        cantidad_existente INTEGER,
        cantidad_comprar INTEGER
    )
    """)

    conn.commit()
    conn.close()

init_db()

# =========================
# CARGAR EXCEL AL INICIAR
# =========================

PRODUCTOS_EXCEL = {}

def cargar_excel():
    global PRODUCTOS_EXCEL
    try:
        ruta = os.path.join(os.getcwd(), "productos.xlsx")
        wb = load_workbook(ruta)
        sheet = wb.active

        for row in sheet.iter_rows(min_row=2, values_only=True):
            codigo, nombre = row
            if codigo is not None:
                PRODUCTOS_EXCEL[str(codigo).strip()] = str(nombre).strip()

        print("Productos cargados:", len(PRODUCTOS_EXCEL))

    except Exception as e:
        print("Error cargando Excel:", e)

cargar_excel()

@app.get("/buscar_producto/{codigo}")
def buscar_producto(codigo: str):
    codigo = codigo.strip()
    nombre = PRODUCTOS_EXCEL.get(codigo, "")
    return {"nombre": nombre}

# =========================
# PRODUCTOS
# =========================

@app.get("/productos")
def obtener_productos():
    conn = sqlite3.connect("database.db")
    cursor = conn.cursor()
    cursor.execute("SELECT * FROM productos")
    datos = cursor.fetchall()
    conn.close()
    return datos

@app.post("/productos")
def agregar_producto(producto: dict):
    conn = sqlite3.connect("database.db")
    cursor = conn.cursor()

    cursor.execute("""
        INSERT INTO productos
        (codigo, nombre, compra, porcentaje, venta, cantidad_existente, cantidad_comprar)
        VALUES (?, ?, ?, ?, ?, ?, ?)
    """, (
        producto["codigo"],
        producto["nombre"],
        producto["compra"],
        producto["porcentaje"],
        producto["venta"],
        producto["cantidad_existente"],
        producto["cantidad_comprar"]
    ))

    conn.commit()
    conn.close()

    return {"mensaje": "Producto agregado"}

@app.delete("/productos/{id}")
def comprar_producto(id: int):
    conn = sqlite3.connect("database.db")
    cursor = conn.cursor()

    cursor.execute("SELECT * FROM productos WHERE id = ?", (id,))
    producto = cursor.fetchone()

    if producto:
        cursor.execute("""
            INSERT INTO historial
            (codigo, nombre, compra, porcentaje, venta, cantidad_existente, cantidad_comprar)
            VALUES (?, ?, ?, ?, ?, ?, ?)
        """, producto[1:])

        cursor.execute("DELETE FROM productos WHERE id = ?", (id,))

    conn.commit()
    conn.close()

    return {"mensaje": "Producto movido al historial"}

@app.get("/historial")
def obtener_historial():
    conn = sqlite3.connect("database.db")
    cursor = conn.cursor()
    cursor.execute("SELECT * FROM historial")
    datos = cursor.fetchall()
    conn.close()
    return datos

# =========================
# FRONTEND
# =========================

app.mount("/static", StaticFiles(directory="static"), name="static")

@app.get("/")
def home():
    return FileResponse("static/index.html")

    conn.close()
    return datos

app.mount("/", StaticFiles(directory="static", html=True), name="static")
